from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook:Workbook = load_workbook(WORKBOOK_PATH)
#nome da planilha                   
sheet_name = "Minha planilha"

#selecionou a planilha 
worksheet: Worksheet = workbook[sheet_name]

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')
        if cell.value == 'Maria':
            worksheet.cell(cell.row,2,23)
    print()
print(worksheet['B3'].value)
    
#workbook.save(WORKBOOK_PATH)