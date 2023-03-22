from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
#nome da planilha 
sheet_name = "Minha planilha"
#crinado planilha 
workbook.create_sheet(sheet_name,0)

#selecionou a planilha 
worksheet: Worksheet = workbook[sheet_name]

#removendo planilha 
workbook.remove(workbook["Sheet"])
#criando cabecalhos 
print(workbook.sheetnames)
worksheet.cell(1,1,'Nome')
worksheet.cell(1,2,'Idade')
worksheet.cell(1,3,'Nota')

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10] 
    ]


# for i,student_row in enumerate(students,start=2):
#     for j, student_collumn in enumerate(student_row,start=1):
#         worksheet.cell(i,j,student_collumn)

#outra maneira de fazer 
for student in students: 
    worksheet.append(students)
    
workbook.save(WORKBOOK_PATH)